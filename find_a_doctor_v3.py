import csv
import re
from bs4 import BeautifulSoup
import requests
import traceback
from geopy.distance import geodesic
import pgeocode
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime

class LekarzValidator:
    def __init__(self, csv_file_path, max_distance_km=100, country='PL'):
        self.csv_file_path = csv_file_path
        self.max_distance_km = max_distance_km
        self.country = country

    def extract_postal_code(self, address):
        postal_code_pattern = r'\b\d{2}-\d{3}\b'
        match = re.search(postal_code_pattern, address)
        return match.group(0) if match else None

    def get_coordinates(self, postal_code):
        nomi = pgeocode.Nominatim(self.country)
        location = nomi.query_postal_code(postal_code)
        if location.empty or location.latitude is None or location.longitude is None:
            return None
        return (location.latitude, location.longitude)

    def are_locations_close(self, postal_code1, postal_code2):
        coords1 = self.get_coordinates(postal_code1)
        coords2 = self.get_coordinates(postal_code2)
        if not coords1 or not coords2:
            raise ValueError("Jedna lub obie lokalizacje nie mogły zostać znalezione.")
        distance = geodesic(coords1, coords2).kilometers
        return distance <= self.max_distance_km

    def read_csv_file(self):
        with open(self.csv_file_path, newline='', encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile, delimiter=',')
            next(csvreader)
            data = []
            for row in csvreader:
                data.append((row[0], row[2].split(" ")[0:2], row[2]))
            return data

    def scrap_ranking_lekarzy(self, imie, nazwisko):
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
        }
        url = f"https://www.rankinglekarzy.pl/lekarze/?q={imie}+{nazwisko}"
        response = requests.get(url, headers=headers)
        soup = BeautifulSoup(response.content, 'html.parser')

        try:
            results = soup.find_all("div", class_="rl-box rl-box--wide rl-box--round-corners rl-search-result cf")
            doctors_info = []
            for r in results:
                profil = r.find_all("a", href=True)[0]['href']
                full_profil_url = f"https://www.rankinglekarzy.pl{profil}"
                name = r.find("span", class_="rl-profile-title__span")
                address = r.find("div", class_="rl-question__doctorlocation")
                if address and name:
                    doctors_info.append((address.text.strip(), name.text.strip(), full_profil_url))
            return doctors_info
        except Exception as e:
            traceback.print_exc()
        return []

    def validate_doctors(self):
        data = self.read_csv_file()
        results = []
        for i, (email, name_parts, personal_data) in enumerate(data):
            imie, nazwisko = name_parts
            input_postal_code = self.extract_postal_code(personal_data)
            if input_postal_code is None:
                results.append((email, imie + ' ' + nazwisko, personal_data, "Zagraniczny adres", None, None))
            else:
                doctors_info = self.scrap_ranking_lekarzy(imie, nazwisko)
                if not doctors_info:
                    results.append((email, imie + ' ' + nazwisko, personal_data, "Brak wyników w rankingu", None, None))
                else:
                    match_found = False
                    for found_address, found_name, profile_url in doctors_info:
                        found_postal_code = self.extract_postal_code(found_address)
                        if not found_postal_code:
                            continue

                        try:
                            if self.are_locations_close(input_postal_code, found_postal_code):
                                results.append((email, imie + ' ' + nazwisko, personal_data, "Może być lekarzem",
                                                profile_url, ', '.join([info[2] for info in doctors_info])))
                                match_found = True
                                break
                        except ValueError as ve:
                            results.append((email, imie + ' ' + nazwisko, personal_data, str(ve), None,
                                            ', '.join([info[2] for info in doctors_info])))
                            match_found = True
                            break

                    if not match_found:
                        results.append((email, imie + ' ' + nazwisko, personal_data, "Za daleko", None,
                                        ', '.join([info[2] for info in doctors_info])))

            print(f"Wiersz {i + 1} został przetworzony")

        return results

    def save_results_to_excel(self, results):
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_path = os.path.join(desktop_path, f'find-a-doctor-{timestamp}.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = "Wyniki"

        headers = ["Imie i nazwisko", "Email", "Adres", "Czy lekarz", "Profil", "Znalezione wyniki"]
        ws.append(headers)

        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        very_light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for result in results:
            email, full_name, address, status, profile_url, found_results = result
            row = [full_name, email, address, status, profile_url, found_results]
            ws.append(row)
            row_idx = ws.max_row

            if status == "Może być lekarzem":
                for cell in ws[row_idx]:
                    cell.fill = light_red_fill
            elif status == "Zagraniczny adres":
                for cell in ws[row_idx]:
                    cell.fill = very_light_gray_fill

        wb.save(file_path)
        return file_path


input_path = input("Podaj ścieżkę pliku wejściowego: ")
validator = LekarzValidator(input_path)
validation_results = validator.validate_doctors()
file_path = validator.save_results_to_excel(validation_results)

print(f"Wyniki zostały zapisane w pliku '{file_path}'")